'''Automate the Quarterly Safety Assessment Report'''

import pyodbc as db
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np

def other(x):
    '''Change any sentences to Other for Analysis purposes'''
    if x.lower() in ('yes', 'no', 'n/a', 'blank'):
        return x
    else:
        return 'Other'

def column_size(sheet):
    '''Dynamically adjust the column sizes in excel sheet'''
    column_widths = []
    for row in sheet:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))+2
            else:
                column_widths += [len(str(cell.value))+2]
    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i+1)].width = column_width

def custom_sum(row):
    return row.sum()

#Connect to MS Access DB
dbpath = input('Enter the full path to the database with the extension at the end:\n')
conn = db.connect(r'Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={}'.format(dbpath))

#Create the main query for MS Access
begdate, enddate =input('Enter date range: (ex: 1/1/2020-2/1/2020)\n').split('-')
begdate, enddate = [x.strip() for x in [begdate, enddate]]
query = 'SELECT * FROM Test_table WHERE [3Review Date] Between #{}# and #{}#;'.format(begdate,enddate)

#Create Dataframe with Pandas and close DB connection
df = pd.read_sql(query, conn)
conn.close()

#Change NaN to 'Blank'
df1 = df.fillna('Blank').copy()

#Get only question columns in proper order
quesdf1 = df1[['P1PerpInformed',
       'P2ChInterviewed', 'P3ChNotIntvw', 'P4MaltrtIntvw',
       'P5AdultNotIntvwWhy', 'P6NonMaltrtIntvw', 'P7CollRel', 'P8SU', 'P9DV',
       'P10PH', 'P11PHConcerns', 'P12MH', 'P13MHConcerns', 'P14Edu',
       'P15EduConcerns', 'P16Removed', 'P17LinkedInv', 'P18InvCaseConf',
       'P19Registry', 'P20Time', 'P21sdmtime', 'P22mentalhealth',
       'P23additionalreports', 'P24diligeff', 'P25Invcasconf', 'FE1SolFocusd',
       'FE2Mapping', 'FE3HarmDanger', 'FE4ChildPersp', 'FE5FamNet',
       'FE6Threequest', 'FE7Consultandinform', 'FE8ConsultInformnxt', 'SA1SA',
       'SA2SAChld', 'SA3Reason', 'SA4SftyId', 'SA5Prot', 'SA6SftyInd',
       'SA7Sfty', 'SA8SftyAgrm', 'SA9SftyAdq', 'SA10Sftytim', 'RA1RAAppr',
       'RA2CNarr', 'RA3OverdAppr', 'RA4FinalDec', 'RA5OverdNarr',
       'SEIEvidSafeCare', 'SE2SafeCarehealsaf', 'SE3Safesleep',
       'SE34referrals', 'SE35referralscaregivr', 'SE36referralsmonitor']]

#Change Comments to Other
quesdf1 = quesdf1.applymap(other).copy()

#Get question names from the column heading
questions = list(quesdf1.columns)

#Create pairs from value count index and its values and make lists for Yes, No, N/A, Blank, Other
yes = []
no = []
na = []
blank = []
other = []
for x in list(quesdf1.columns):
    name = list(quesdf1[x].value_counts().index)
    value = list(quesdf1[x].value_counts().values)
    pair = list(zip(name, value))

    #Create a dictionary with the key value
    d = {}
    for key, val in pair:
        d[key] = val

    #Check value and add to correct list
    if 'Yes' in d.keys():
        yes.append(d['Yes'])
    else:
        yes.append(0)
        
    if 'No' in d.keys():
        no.append(d['No'])
    else:
        no.append(0)
        
    if 'N/A' in d.keys():
        na.append(d['N/A'])
    else:
        na.append(0)
        
    if 'Blank' in d.keys():
        blank.append(d['Blank'])
    else:
        blank.append(0)
        
    if 'Other' in d.keys():
        other.append(d['Other'])
    else:
        other.append(0)
    #print(d)
    #print('Sum:'+str(sum(d.values())))

#Create df for analysis and add %

adf = pd.DataFrame(list(zip(questions, yes, no, na, blank, other)), columns = ['Questions', 'Yes', 'No', 'N/A', 'Blank', 'Other'])

total = sum(d.values())

adf.insert(2, 'Yes %', (adf['Yes']/total), False)
adf.insert(4, 'No %', (adf['No']/total), False)
adf.insert(6, 'N/A %', (adf['N/A']/total), False)
adf.insert(8, 'Blank %', (adf['Blank']/total), False)
adf['Other %'] = (adf['Other']/total) 
adf['Total'] = total

#Create pivot tables for the Safety Assessment questions
sas = ['SA1SA', 'SA2SAChld', 'SA3Reason', 'SA4SftyId',
       'SA5Prot', 'SA6SftyInd', 'SA7Sfty', 'SA8SftyAgrm', 'SA9SftyAdq',
       'SA10Sftytim']

tables = []

for sa in sas:
    columns = ['Yes', 'No','N/A', 'Blank']
    regions = ['Beech Street', 'Kent Co.', 'Sussex Co.', 'UP']
    table = pd.pivot_table(data = df1, index=['7Region'], columns=[sa], aggfunc = {sa:'count'}, fill_value = 0).sort_index(axis=1, ascending=False)
    table = table[sa].copy()

    for col in columns:             #Check if table has all the columns to easily add later for a totals table
        if col in table.columns:
            continue
        else:
            table[col] = 0

    for idx in regions:
        if idx in table.index:
            continue
        else:
            table.loc[idx] = 0

    table = table.sort_index(axis=0, ascending = True)
    table = table.sort_index(axis=1, ascending = False)
    table['Total'] = table.apply(custom_sum, axis=1)
    table.loc['Total'] = table.apply(custom_sum, axis = 0)
    table['% in Conformity'] = np.where(table['Yes'] == 0, 0, (table['Yes']/(table['Yes']+table['No'])))
    table.index.name = 'Region'
    tables.append(table)

total_table = pd.DataFrame()        #Create a total tables to add all the tables together
for table in tables:
    if total_table.empty:
        total_table = table.copy()
    else:
        total_table += table.copy()

#total_table.columns.name = 'Region'
#total_table.index.name = None

total_table['% in Conformity'] = np.where(total_table['Yes'] == 0, 0, (total_table['Yes']/(total_table['Yes']+total_table['No'])))



#writer = pd.ExcelWriter('tables.xlsx', engine = 'xlsxwriter', datetime_format = 'mm/dd/yyyy')
  
#Export Question Analysis to excel and make data into a table
path = input('Where do you want to save the file? (ex: C:\\User\\Folder\\filename.xlsx)\n')
writer = pd.ExcelWriter(path, engine = 'xlsxwriter', datetime_format = 'mm/dd/yyyy')
adf.to_excel(writer, sheet_name = 'Safety Analysis', startrow=1, header=False, index=False)

column_settings = [{'header': column} for column in adf.columns]
(max_row, max_col) = adf.shape

workbook = writer.book
worksheet = writer.sheets['Safety Analysis']
worksheet.add_table(0,0, max_row, max_col - 1, {'columns' : column_settings, 'style': 'Table Style Light 9'})

percent_fmt = workbook.add_format({'num_format':'0.00%'}) #Adjust column width and add formatting
worksheet.set_column('A:A', 20)
worksheet.set_column('C:C', None, percent_fmt)
worksheet.set_column('E:E', None, percent_fmt)
worksheet.set_column('G:G', None, percent_fmt)
worksheet.set_column('I:I', None, percent_fmt)
worksheet.set_column('K:K', None, percent_fmt)

#Export Safety Assessment tables to excel and put on same sheet with formatting and headers
xrows =list(range(3,103,10))
xhead = [
    'SA1. Was the Safety Assessment completed on the appropriate household(s)?',
    'SA2. Was safety assessed for all children in the household?',
    'SA3. If "No" to Question SA2, was the reason documented?',
    'SA4. Were all safety threats identified for each child?',
    'SA5. Were the identified protective capacities documented during the contact(s) with the family?',
    'SA6. Were the indicated safety interventions appropriate for the identified threats?',
    'SA7. Is the final safety finding correct/appropriate?',
    'SA8. Was a Child Safety Agreement completed according to policy?',
    'SA9. If a Child Safety Agreement was completed, did it address the threats adequately?',
    'SA10 If a Child Safety Agreement was completed, was the Child S'
]

for idx in range(len(tables)):          #Put each table into same sheet and add question name before table with formattting
    tables[idx].to_excel(writer, sheet_name='Safety', float_format="%.4f", startrow=xrows[idx]+1, header=False)
    wb = writer.book
    bold  = wb.add_format({'bold': True, 'bg_color': '#4f81bd', 'font_color':'white'})
    sheet = writer.sheets['Safety']
    (max_row, max_col)=table.shape
    sheet.add_table(xrows[idx], 0, xrows[idx]+ max_row, max_col, {'columns':[{'header': 'Region'}, {'header': 'Yes'}, {'header': 'No'}, {'header': 'N/A'}, {'header': 'Blank'}, {'header': 'Total'},{'header': '% in Conformity'}], 'style':'Table Style Light 9'}) 
    #sheet.write(xrows[idx]-2, 0, xhead[idx])
    sheet.merge_range(xrows[idx]-2, 0,xrows[idx]-2,6,xhead[idx] ,bold)
    
total_row = xrows[-1]+10
total_table.to_excel(writer, sheet_name='Safety', float_format="%.4f", startrow=total_row+1, header=False)
(max_row, max_col) = total_table.shape
sheet.add_table(total_row, 0, total_row + max_row, max_col, {'columns':[{'header': 'Region'}, {'header': 'Yes'}, {'header': 'No'}, {'header': 'N/A'}, {'header': 'Blank'}, {'header': 'Total'},{'header': '% in Conformity'}], 'style':'Table Style Medium 1'})
total_fmt = wb.add_format({'bold':True, 'bg_color': 'black', 'font_color':'white'})
sheet.merge_range(total_row - 2, 0, total_row - 2, 6, 'COMBINED SAFETY ASSESSMENT', total_fmt)

sheet.set_column('A:A', 20)
sheet.set_column('G:G', 15, percent_fmt)
#writer.save()

#Export Raw Data to Excel
df1.to_excel(writer, sheet_name = 'Raw Data', index=None)      

writer.save()


#Adjust column sizes in Raw Data Tab

book = load_workbook(path)
worksheet = book['Raw Data']
column_size(worksheet)
book.save(path)
#for sheet in ['Safety Analysis', 'Raw Data']:
#    worksheet = book[sheet]
#    column_size(worksheet)
#book.save(path)
